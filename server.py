from __future__ import annotations

import json
import hashlib
import mimetypes
import os
import queue
import re
import shutil
import threading
import time
import urllib.error
import urllib.request
from collections import defaultdict
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlencode, unquote, urlparse
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent


def load_env_file(path: Path) -> None:
    """读取本地环境变量文件，API Key 只进入后端进程环境，不输出到前端。"""
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


load_env_file(ROOT / ".env")
load_env_file(ROOT / "apifootball.env")

DAILY_DIR = ROOT / "每日更新"
UPLOAD_DIR = DAILY_DIR / "Excel表格放这里"
IMPORTED_DIR = UPLOAD_DIR / "已导入"
FAILED_DIR = UPLOAD_DIR / "导入失败"
DATA_DIR = ROOT / "data"
DATA_FILE = DATA_DIR / "schedule.json"
TEAM_ALIASES_FILE = DATA_DIR / "team_aliases.json"
LEAGUE_ALIASES_FILE = DATA_DIR / "league_aliases.json"
API_USAGE_FILE = DATA_DIR / "api_usage.json"
API_LOG_FILE = DATA_DIR / "apifootball_api.log"
AI_REPORTS_FILE = DATA_DIR / "ai_reports.json"
TEAM_LOGO_CACHE_DIR = DATA_DIR / "team-logo-cache"
DAILY_IMAGE_DIRS = {
    "draw": DAILY_DIR / "今日开奖图片放这里",
    "homework": DAILY_DIR / "作业图片放这里",
    "contact": DAILY_DIR / "联系店主图片放这里",
}
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
TEAM_LOGO_EXTENSIONS = {".svg", ".png", ".jpg", ".jpeg", ".webp"}
TEAM_LOGO_ALLOWED_HOSTS = {
    "jsd.onmicrosoft.cn",
    "flagcdn.com",
    "purecatamphetamine.github.io",
    "cdn.prod.website-files.com",
    "logotyp.us",
    "logotypes.dev",
}
TEAM_LOGO_MAX_BYTES = 2 * 1024 * 1024
HOST = "0.0.0.0"
PORT = 5180
SCAN_SECONDS = 2
RETENTION_DAYS = 7
VISIBLE_PAST_DAYS = 5
API_FOOTBALL_BASE_URL = "https://v3.football.api-sports.io"
API_FOOTBALL_API_KEY = os.environ.get("API_FOOTBALL_API_KEY", "")
API_DAILY_LIMIT = int(os.environ.get("API_DAILY_LIMIT", "100") or "100")
ADMIN_TOKEN = os.environ.get("ADMIN_TOKEN", "")
DEEPSEEK_API_KEY = os.environ.get("DEEPSEEK_API_KEY", "")
DOUBAO_API_KEY = os.environ.get("DOUBAO_API_KEY", "")
DOUBAO_ENDPOINT_ID = os.environ.get("DOUBAO_ENDPOINT_ID", "")
LOCAL_TIMEZONE = "Asia/Shanghai"
SCORE_POLL_MIN_REMAINING = 25
DATE_STRIP_OFFSETS = (-4, -3, -2, -1, 0, 1)

CLIENTS: list[queue.Queue[dict]] = []
CLIENTS_LOCK = threading.Lock()
DATA_LOCK = threading.Lock()
IN_PROGRESS: set[Path] = set()
PROCESSED_UPLOAD_SIGNATURES: dict[str, tuple[int, int]] = {}

HEADER_ALIASES = {
    "match_no": ["竞彩编号", "编号", "赛事编号", "场次编号", "竞彩", "match_no", "matchno", "id"],
    "league": ["赛事种类", "赛事", "联赛", "赛事名称", "比赛类型", "league"],
    "teams": ["双方队伍", "对阵", "主客队", "比赛球队", "队伍", "双方球队", "teams", "match"],
    "home": ["主队", "主场", "主场队伍", "home"],
    "away": ["客队", "客场", "客场队伍", "away"],
    "kickoff": ["开赛日期", "开赛时间", "比赛时间", "时间", "日期", "kickoff", "time"],
    "round": ["轮次", "轮", "第几轮", "round"],
}

TEAM_SEPARATORS = [
    r"\s*vs\s*",
    r"\s*v\s*",
    r"\s*对\s*",
    r"\s*—\s*",
    r"\s*–\s*",
    r"\s+-\s+",
]

BADGES = ["🔴", "🔵", "🟢", "🟡", "🟠", "🟣", "⚪", "⚫"]
WEEKDAY_INDEX = {
    "周一": 0,
    "周二": 1,
    "周三": 2,
    "周四": 3,
    "周五": 4,
    "周六": 5,
    "周日": 6,
    "周天": 6,
}

DEFAULT_TEAM_ALIASES = {
    "中国": ["China", "China PR"],
    "中国香港": ["Hong Kong", "Hong Kong China"],
    "中国澳门": ["Macau", "Macao"],
    "中华台北": ["Chinese Taipei", "Taiwan"],
    "日本": ["Japan"],
    "韩国": ["South Korea", "Korea Republic"],
    "朝鲜": ["North Korea", "Korea DPR"],
    "澳大利亚": ["Australia"],
    "新西兰": ["New Zealand"],
    "新加坡": ["Singapore"],
    "泰国": ["Thailand"],
    "越南": ["Vietnam"],
    "马来西亚": ["Malaysia"],
    "印度尼西亚": ["Indonesia"],
    "印尼": ["Indonesia"],
    "菲律宾": ["Philippines"],
    "印度": ["India"],
    "伊朗": ["Iran"],
    "伊拉克": ["Iraq"],
    "沙特": ["Saudi Arabia"],
    "沙特阿拉伯": ["Saudi Arabia"],
    "卡塔尔": ["Qatar"],
    "阿联酋": ["United Arab Emirates", "UAE"],
    "阿曼": ["Oman"],
    "约旦": ["Jordan"],
    "乌兹别克": ["Uzbekistan"],
    "阿根廷": ["Argentina"],
    "巴西": ["Brazil"],
    "乌拉圭": ["Uruguay"],
    "智利": ["Chile"],
    "哥伦比亚": ["Colombia"],
    "英格兰": ["England"],
    "法国": ["France"],
    "德国": ["Germany"],
    "意大利": ["Italy"],
    "西班牙": ["Spain"],
    "葡萄牙": ["Portugal"],
    "荷兰": ["Netherlands"],
    "比利时": ["Belgium"],
    "瑞士": ["Switzerland"],
    "丹麦": ["Denmark"],
    "乌克兰": ["Ukraine"],
    "瑞典": ["Sweden"],
    "挪威": ["Norway"],
    "芬兰": ["Finland"],
    "奥地利": ["Austria"],
    "捷克": ["Czech Republic", "Czechia"],
    "斯洛伐克": ["Slovakia"],
    "斯洛文尼亚": ["Slovenia"],
    "克罗地亚": ["Croatia"],
    "塞尔维亚": ["Serbia"],
    "黑山": ["Montenegro"],
    "匈牙利": ["Hungary"],
    "土耳其": ["Turkey"],
    "希腊": ["Greece"],
    "加拿大": ["Canada"],
    "爱尔兰": ["Ireland"],
    "北爱尔兰": ["N.Ireland"],
    "秘鲁": ["Peru"],
    "苏格兰": ["Scotland"],
    "威尔士": ["Wales"],
    "美国": ["United States", "USA"],
    "墨西哥": ["Mexico"],
    "突尼斯": ["Tunisia"],
    "埃及": ["Egypt"],
    "波黑": ["Bosnia-Herzegovina", "Bosnia"],
    "玻利维亚": ["Bolivia"],
    "委内瑞拉": ["Venezuela"],
    "洪都拉斯": ["Honduras"],
    "巴拿马": ["Panama"],
    "罗马尼亚": ["Romania"],
    "摩洛哥": ["Morocco"],
    "横滨水手": ["Yokohama F. Marinos", "Yokohama Marinos"],
    "神户胜利": ["Vissel Kobe"],
    "神户胜利船": ["Vissel Kobe"],
    "浦和红钻": ["Urawa Red Diamonds", "Urawa Reds"],
    "鹿岛鹿角": ["Kashima Antlers"],
    "町田泽维": ["Machida Zelvia", "FC Machida Zelvia"],
    "名古屋鲸": ["Nagoya Grampus"],
    "冈山绿雉": ["Fagiano Okayama"],
    "清水鼓动": ["Shimizu S-Pulse", "Shimizu S-Pulse"],
    "柏太阳神": ["Kashiwa Reysol"],
    "川崎前锋": ["Kawasaki Frontale"],
    "广岛三箭": ["Sanfrecce Hiroshima"],
    "大阪钢巴": ["Gamba Osaka"],
    "大阪樱花": ["Cerezo Osaka"],
    "京都": ["Kyoto Sanga", "Kyoto Sanga FC"],
    "东京绿茵": ["Tokyo Verdy"],
    "FC东京": ["FC Tokyo"],
    "福冈": ["Avispa Fukuoka"],
    "新泻天鹅": ["Albirex Niigata"],
    "札幌冈萨多": ["Consadole Sapporo", "Hokkaido Consadole Sapporo"],
    "蔚山HD": ["Ulsan HD", "Ulsan Hyundai"],
    "全北现代": ["Jeonbuk Hyundai Motors"],
    "浦项制铁": ["Pohang Steelers"],
    "首尔FC": ["FC Seoul"],
    "水原三星": ["Suwon Samsung Bluewings"],
    "曼城": ["Manchester City", "Man City"],
    "曼联": ["Manchester United", "Man United"],
    "利物浦": ["Liverpool"],
    "切尔西": ["Chelsea"],
    "阿森纳": ["Arsenal"],
    "热刺": ["Tottenham Hotspur", "Spurs"],
    "皇马": ["Real Madrid"],
    "皇家马德里": ["Real Madrid"],
    "巴萨": ["Barcelona", "FC Barcelona"],
    "巴塞罗那": ["Barcelona", "FC Barcelona"],
    "马竞": ["Atletico Madrid", "Atlético Madrid"],
    "拜仁": ["Bayern Munich", "FC Bayern Munich"],
    "多特": ["Borussia Dortmund", "Dortmund"],
    "巴黎圣日耳曼": ["Paris Saint-Germain", "PSG"],
    "国际米兰": ["Inter Milan", "Internazionale"],
    "AC米兰": ["AC Milan"],
    "尤文": ["Juventus"],
    "罗马": ["Roma", "AS Roma"],
    "那不勒斯": ["Napoli", "SSC Napoli"],
}

DEFAULT_LEAGUE_ALIASES = {
    "日职": ["J1 League", "Japan J1 League"],
    "日乙": ["J2 League", "Japan J2 League"],
    "日联赛杯": ["J-League Cup", "J. League Cup"],
    "韩职": ["K League 1"],
    "中超": ["Chinese Super League"],
    "英超": ["Premier League"],
    "英冠": ["Championship"],
    "西甲": ["La Liga"],
    "意甲": ["Serie A"],
    "德甲": ["Bundesliga"],
    "法甲": ["Ligue 1"],
    "欧冠": ["UEFA Champions League", "Champions League"],
    "欧联": ["UEFA Europa League", "Europa League"],
    "欧协联": ["UEFA Europa Conference League", "Conference League"],
    "国际赛": ["Friendlies", "International Friendlies", "International Friendly", "World Friendlies", "World Friendly", "International", "Friendly", "World"],
    "世界杯": ["World Cup", "FIFA World Cup"],
    "欧洲杯": ["European Championship", "UEFA Euro"],
    "美职": ["Major League Soccer", "MLS"],
}


def now_iso() -> str:
    return datetime.now().replace(microsecond=0).isoformat()


def safe_print(message: str) -> None:
    try:
        print(message, flush=True)
    except Exception:
        pass


def ensure_dirs() -> None:
    for folder in (UPLOAD_DIR, IMPORTED_DIR, FAILED_DIR, DATA_DIR, TEAM_LOGO_CACHE_DIR):
        folder.mkdir(parents=True, exist_ok=True)

    ensure_alias_file(TEAM_ALIASES_FILE, DEFAULT_TEAM_ALIASES)
    ensure_alias_file(LEAGUE_ALIASES_FILE, DEFAULT_LEAGUE_ALIASES)
    ensure_api_usage_file()
    cleanup_old_upload_archives()

    for folder in DAILY_IMAGE_DIRS.values():
        folder.mkdir(parents=True, exist_ok=True)
        note = folder / "请把图片放这里.txt"
        if not note.exists():
            note.write_text(
                "把要展示的图片放到这个文件夹，网站会自动读取最新一张图片。支持 jpg、jpeg、png、webp、gif。\n",
                encoding="utf-8",
            )

    note = UPLOAD_DIR / "请把表格放这里.txt"
    if not note.exists():
        note.write_text(
            "把每日 Excel 放到这个文件夹后，网页会自动导入并刷新。\n"
            "支持赛程表头：竞彩编号、赛事种类、双方队伍、开赛日期。\n"
            "导入成功的文件会移动到“已导入”，失败文件会移动到“导入失败”。\n",
            encoding="utf-8",
        )


def cleanup_old_upload_archives() -> None:
    cutoff = time.time() - RETENTION_DAYS * 24 * 60 * 60
    for folder in (IMPORTED_DIR, FAILED_DIR):
        if not folder.exists():
            continue
        for path in folder.iterdir():
            try:
                if path.is_file() and path.stat().st_mtime < cutoff:
                    path.unlink()
            except OSError:
                pass


def empty_data() -> dict:
    return {"matches": [], "lastImport": None}


def load_data() -> dict:
    if not DATA_FILE.exists():
        return empty_data()
    try:
        data = json.loads(DATA_FILE.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return empty_data()
    if "matches" not in data:
        data["matches"] = []
    if "lastImport" not in data:
        data["lastImport"] = None
    for match in data.get("matches", []):
        ensure_match_fields(match)
    return data


def save_data(data: dict) -> None:
    tmp = DATA_FILE.with_suffix(DATA_FILE.suffix + ".tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(DATA_FILE)


def read_json_file(path: Path, fallback):
    if not path.exists():
        return fallback
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return fallback


def write_json_file(path: Path, payload) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def ensure_alias_file(path: Path, defaults: dict) -> None:
    if path.exists():
        return
    write_json_file(path, defaults)


def today_key() -> str:
    return date.today().isoformat()


def local_today() -> date:
    return datetime.now(ZoneInfo(LOCAL_TIMEZONE)).date()


def empty_api_usage() -> dict:
    return {
        "date": today_key(),
        "used": 0,
        "remaining": API_DAILY_LIMIT,
        "limit": API_DAILY_LIMIT,
        "lastUpdatedAt": None,
        "lastRequestAt": None,
        "lastSuccessAt": None,
        "lastError": "",
        "consecutiveFailures": 0,
        "pauseUntil": None,
        "nextAutoUpdateAt": None,
    }


def normalize_api_usage(usage: dict) -> dict:
    if usage.get("date") != today_key():
        return empty_api_usage()
    usage.setdefault("used", 0)
    usage.setdefault("remaining", max(0, API_DAILY_LIMIT - int(usage.get("used", 0))))
    usage.setdefault("limit", API_DAILY_LIMIT)
    usage.setdefault("lastUpdatedAt", None)
    usage.setdefault("lastRequestAt", None)
    usage.setdefault("lastSuccessAt", None)
    usage.setdefault("lastError", "")
    usage.setdefault("consecutiveFailures", 0)
    usage.setdefault("pauseUntil", None)
    usage.setdefault("nextAutoUpdateAt", None)
    return usage


def load_api_usage() -> dict:
    return normalize_api_usage(read_json_file(API_USAGE_FILE, empty_api_usage()))


def save_api_usage(usage: dict) -> None:
    write_json_file(API_USAGE_FILE, normalize_api_usage(usage))


def ensure_api_usage_file() -> None:
    if not API_USAGE_FILE.exists():
        save_api_usage(empty_api_usage())


def append_api_log(entry: dict) -> None:
    entry.setdefault("loggedAt", now_iso())
    API_LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    with API_LOG_FILE.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(entry, ensure_ascii=False) + "\n")


def load_aliases(path: Path, defaults: dict) -> dict:
    aliases = dict(defaults)
    custom = read_json_file(path, {})
    if isinstance(custom, dict):
        for key, values in custom.items():
            if isinstance(values, list):
                merged = list(dict.fromkeys([*aliases.get(key, []), *[str(item) for item in values]]))
                aliases[str(key)] = merged
            elif values:
                aliases[str(key)] = list(dict.fromkeys([*aliases.get(key, []), str(values)]))
    return aliases


def normalize_header(value) -> str:
    text = str(value or "").strip()
    text = re.sub(r"[\s:：/\\_\-（）()]+", "", text)
    return text.lower()


def normalize_team(value) -> str:
    return re.sub(r"\s+", "", str(value or "").strip())


def cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if text.endswith(".0"):
        return text[:-2]
    return text


def resolve_headers(raw_headers: list) -> dict[str, int]:
    normalized = [normalize_header(item) for item in raw_headers]
    resolved: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_key = normalize_header(alias)
            if alias_key in normalized:
                resolved[field] = normalized.index(alias_key)
                break
    return resolved


def split_teams(value: str) -> tuple[str, str]:
    text = str(value or "").strip()
    text = re.sub(r"\s+", " ", text)
    for pattern in TEAM_SEPARATORS:
        parts = re.split(pattern, text, maxsplit=1, flags=re.IGNORECASE)
        if len(parts) == 2 and parts[0].strip() and parts[1].strip():
            return parts[0].strip(), parts[1].strip()

    compact = re.split(r"\s{2,}", str(value or "").strip(), maxsplit=1)
    if len(compact) == 2 and compact[0].strip() and compact[1].strip():
        return compact[0].strip(), compact[1].strip()

    raise ValueError(f"无法拆分双方队伍：{value}")


def get_row_teams(row: dict) -> tuple[str, str]:
    if row.get("teams"):
        return split_teams(row.get("teams", ""))
    home = str(row.get("home", "")).strip()
    away = str(row.get("away", "")).strip()
    if home and away:
        return home, away
    raise ValueError("缺少双方队伍，或缺少主队/客队")


def parse_score(value: str) -> tuple[str, str, str]:
    text = str(value or "").strip()
    match = re.search(r"(\d+)\D+(\d+)", text)
    if not match:
        raise ValueError(f"比分格式无法识别：{value}")
    home, away = match.group(1), match.group(2)
    return home, away, f"{home}∶{away}"


def parse_score_safe(value) -> tuple[str, str, str] | None:
    try:
        return parse_score(str(value or ""))
    except ValueError:
        return None


def display_score_from_fields(match: dict) -> str:
    return str(match.get("apiFullTimeScore") or match.get("manualFullTimeScore") or match.get("currentScore") or "")


def display_half_score_from_fields(match: dict) -> str:
    return str(match.get("apiHalfTimeScore") or match.get("manualHalfTimeScore") or "")


def apply_effective_score_fields(match: dict) -> None:
    full_score = display_score_from_fields(match)
    half_score = display_half_score_from_fields(match)
    parsed = parse_score_safe(full_score)
    match["fullScore"] = full_score
    match["halfScore"] = half_score
    if parsed:
        match["homeScore"], match["awayScore"], _ = parsed
    elif full_score:
        match["homeScore"] = ""
        match["awayScore"] = ""
        match["needsManualCheck"] = True
    else:
        match["homeScore"] = ""
        match["awayScore"] = ""


def ensure_match_fields(match: dict) -> dict:
    match.setdefault("apiMatchId", "")
    match.setdefault("homeTeamAliasMatched", "")
    match.setdefault("awayTeamAliasMatched", "")
    match.setdefault("leagueAliasMatched", "")
    match.setdefault("matchStatus", "未开始")
    match.setdefault("currentScore", "")
    match.setdefault("penaltyScore", "")
    match.setdefault("extraTimeScore", "")
    match.setdefault("scoreSource", "pending")
    match.setdefault("manualHalfTimeScore", match.get("halfScore", ""))
    match.setdefault("manualFullTimeScore", match.get("fullScore", ""))
    match.setdefault("apiHalfTimeScore", "")
    match.setdefault("apiFullTimeScore", "")
    match.setdefault("manualLocked", False)
    match.setdefault("needsManualCheck", False)
    match.setdefault("lastApiState", "")
    match.setdefault("lastUpdatedAt", match.get("updatedAt", ""))
    match.setdefault("apiUpdatedAt", "")
    match.setdefault("manualUpdatedAt", "")
    if not match.get("kickoffDisplay") and match.get("kickoff"):
        kickoff = kickoff_datetime(match)
        match["kickoffDisplay"] = kickoff.strftime("%Y-%m-%d %H:%M") if kickoff else str(match.get("kickoff") or "")
    apply_effective_score_fields(match)
    return match


def parse_kickoff(value) -> datetime:
    if isinstance(value, datetime):
        return value.replace(second=0, microsecond=0)

    text = cell_text(value)
    text = text.replace("：", ":").replace("年", "-").replace("月", "-").replace("日", "")
    text = re.sub(r"\s+", " ", text).strip()

    formats = [
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%m-%d %H:%M",
        "%m/%d %H:%M",
    ]

    for fmt in formats:
        try:
            parsed = datetime.strptime(text, fmt)
            if "%Y" not in fmt:
                parsed = parsed.replace(year=datetime.now().year)
            return parsed.replace(second=0, microsecond=0)
        except ValueError:
            continue

    match = re.search(r"(?:(\d{4})[-/])?(\d{1,2})[-/](\d{1,2})\s+(\d{1,2}):(\d{2})", text)
    if match:
        year = int(match.group(1) or datetime.now().year)
        month = int(match.group(2))
        day = int(match.group(3))
        hour = int(match.group(4))
        minute = int(match.group(5))
        return datetime(year, month, day, hour, minute)

    raise ValueError(f"开赛日期格式无法识别：{value}")


def day_payload(kickoff: datetime) -> dict:
    return {
        "dateKey": kickoff.strftime("%Y-%m-%d"),
        "dateLabel": f"{kickoff.month}月{kickoff.day}日",
        "dayNumber": str(kickoff.day),
        "dayOfWeek": ["周一", "周二", "周三", "周四", "周五", "周六", "周日"][kickoff.weekday()],
        "time": kickoff.strftime("%H:%M"),
    }


def schedule_day_for_match(match_no: str, kickoff: datetime) -> datetime:
    """竞彩编号决定日期栏归属，Excel 开赛日期只保存在 kickoff 中。"""
    text = normalize_match_no(match_no)
    weekday_match = re.search(r"(周[一二三四五六日天])", text)
    if not weekday_match:
        return kickoff

    target_weekday = WEEKDAY_INDEX.get(weekday_match.group(1))
    if target_weekday is None:
        return kickoff

    today = local_today()
    for offset in DATE_STRIP_OFFSETS:
        candidate = today + timedelta(days=offset)
        if candidate.weekday() == target_weekday:
            return datetime.combine(candidate, kickoff.time())
    return kickoff


def normalize_match_no(value: str) -> str:
    text = re.sub(r"\s+", "", str(value or "").strip())
    weekday_match = re.search(r"(周[一二三四五六日天])", text)
    number_match = re.search(r"(\d{1,3})\D*$", text)
    if weekday_match and number_match:
        weekday = weekday_match.group(1).replace("周天", "周日")
        return f"{weekday}{int(number_match.group(1)):03d}"
    if number_match:
        return f"{int(number_match.group(1)):03d}"
    return text


def team_badge(team: str) -> str:
    return BADGES[sum(ord(char) for char in team) % len(BADGES)]


def read_excel(path: Path) -> tuple[str, list[dict], dict[str, int]]:
    if path.suffix.lower() == ".xls":
        raise ValueError("当前环境不支持老版 .xls，请另存为 .xlsx 后再上传。")

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        workbook_type = ""
        workbook_records: list[dict] = []
        workbook_headers: dict[str, int] = {}
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            for header_index, row in enumerate(rows):
                if not row or sum(1 for cell in row if cell not in (None, "")) < 2:
                    continue

                headers = list(row)
                resolved = resolve_headers(headers)
                has_teams = "teams" in resolved or all(key in resolved for key in ("home", "away"))
                is_schedule = all(key in resolved for key in ("match_no", "league", "kickoff")) and has_teams
                if not is_schedule:
                    continue

                table_type = "schedule"
                if workbook_type and table_type != workbook_type:
                    continue

                records = []
                for source_row, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
                    if not values or not any(cell not in (None, "") for cell in values):
                        continue
                    item = {"__row": source_row}
                    for field, column in resolved.items():
                        item[field] = cell_text(values[column] if column < len(values) else "")
                    records.append(item)

                workbook_type = table_type
                workbook_headers = resolved
                workbook_records.extend(records)
        if workbook_type:
            return workbook_type, workbook_records, workbook_headers
    finally:
        workbook.close()

    raise ValueError("未找到可识别表头。赛程表需包含：竞彩编号、赛事种类、双方队伍、开赛日期。")


def purge_old_matches(data: dict) -> int:
    cutoff = date.today() - timedelta(days=VISIBLE_PAST_DAYS)
    before = len(data.get("matches", []))
    data["matches"] = [
        match for match in data.get("matches", [])
        if date.fromisoformat(match["dateKey"]) >= cutoff
    ]
    return before - len(data["matches"])


def load_ai_reports() -> dict:
    if not AI_REPORTS_FILE.exists():
        return {}
    try:
        return json.loads(AI_REPORTS_FILE.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}

def save_ai_reports(reports: dict) -> None:
    AI_REPORTS_FILE.write_text(json.dumps(reports, ensure_ascii=False, indent=2), encoding="utf-8")

def purge_old_ai_reports(reports: dict) -> dict:
    cutoff = datetime.now(ZoneInfo(LOCAL_TIMEZONE)) - timedelta(days=8)
    cutoff_str = cutoff.isoformat()
    purged = 0
    for match_no in list(reports.keys()):
        entry = reports[match_no]
        if isinstance(entry, dict):
            ages = []
            for ai in ("deepseek", "doubao"):
                t = entry.get(ai, {}).get("generatedAt", "")
                if t:
                    ages.append(t)
            if ages and all(a < cutoff_str for a in ages):
                del reports[match_no]
                purged += 1
            else:
                for ai in list(entry.keys()):
                    t = entry[ai].get("generatedAt", "")
                    if t and t < cutoff_str:
                        del entry[ai]
    return reports

AI_PROMPT = """你是专业足球数据分析师。在报告中适当增加图标和彩色标题。请按照以下模板给出总共600字左右的分析报告：

1. 两队的排名：如果是两支国家队之间，请分别写明主队和客队的FIFA排名；如果是联赛，请写明当前积分赛的排名。
2. 两队近五场历史交锋战绩，并分析对战优劣势。
3. 两支球队近十场比赛的胜负，并分析原因。
4. 双方核心球员的伤病停赛情况，以及其他特殊因素。
5. 务必给出具体的比分预测，和半场是否进球（可以多写几个），并给出具体原因。

请表明自己使用的是哪一个大模型，并最后做出风险提示：赛事存在临场变数，AI分析仅数据参考，不构成投注建议；购彩有风险，理性娱乐、量力而行，未满18周岁禁止购彩。"""

def build_ai_match_context(match: dict) -> str:
    ctx = []
    ctx.append(f"主队: {match.get('home', '')}")
    ctx.append(f"客队: {match.get('away', '')}")
    ctx.append(f"赛事: {match.get('league', '')}")
    kt = match.get("kickoff", "")
    if kt:
        ctx.append(f"开赛时间: {kt}")
    full = match.get("fullScore") or ""
    half = match.get("halfScore") or match.get("apiHalfTimeScore") or ""
    if full:
        ctx.append(f"全场比分: {full}")
    if half:
        ctx.append(f"半场比分: {half}")
    return "\n".join(ctx)

def call_deepseek(match_context: str) -> str:
    if not DEEPSEEK_API_KEY:
        raise ValueError("DeepSeek API Key 未配置")
    body = {
        "model": "deepseek-chat",
        "messages": [
            {"role": "system", "content": AI_PROMPT},
            {"role": "user", "content": f"请分析以下比赛：\n{match_context}"}
        ],
        "temperature": 0.7,
        "max_tokens": 2000,
        "stream": False,
    }
    body_bytes = json.dumps(body).encode("utf-8")
    req = urllib.request.Request(
        "https://api.deepseek.com/v1/chat/completions",
        data=body_bytes,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=120) as resp:
        result = json.loads(resp.read().decode("utf-8"))
    return result["choices"][0]["message"]["content"]

def call_doubao(match_context: str) -> str:
    if not DOUBAO_API_KEY:
        raise ValueError("豆包 API Key 未配置")
    if not DOUBAO_ENDPOINT_ID:
        raise ValueError("豆包 Endpoint ID 未配置（请在 .env 中设置 DOUBAO_ENDPOINT_ID=ep-xxx）")
    body = {
        "model": DOUBAO_ENDPOINT_ID,
        "messages": [
            {"role": "system", "content": AI_PROMPT},
            {"role": "user", "content": f"请分析以下比赛：\n{match_context}"}
        ],
        "temperature": 0.7,
        "max_tokens": 2000,
        "stream": False,
        "enable_search": True,
    }
    body_bytes = json.dumps(body).encode("utf-8")
    req = urllib.request.Request(
        "https://ark.cn-beijing.volces.com/api/v3/chat/completions",
        data=body_bytes,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {DOUBAO_API_KEY}",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            result = json.loads(resp.read().decode("utf-8"))
        return result["choices"][0]["message"]["content"]
    except urllib.error.HTTPError as e:
        body_text = ""
        try:
            body_text = e.read().decode("utf-8")[:500]
        except Exception:
            pass
        raise ValueError(f"HTTP {e.code}: {body_text}")

def generate_ai_report_for_match(match: dict) -> dict:
    match_no = match.get("matchNo", "")
    if not match_no:
        return {"matchNo": match_no, "reports": {}}

    ctx = build_ai_match_context(match)
    reports: dict[str, dict] = {}
    errors = []

    if DEEPSEEK_API_KEY:
        try:
            content = call_deepseek(ctx)
            reports["deepseek"] = {
                "content": content,
                "generatedAt": now_iso(),
            }
        except Exception as exc:
            errors.append(f"DeepSeek: {exc}")
    else:
        errors.append("DeepSeek API Key 未配置")

    if DOUBAO_API_KEY and DOUBAO_ENDPOINT_ID:
        try:
            content = call_doubao(ctx)
            reports["doubao"] = {
                "content": content,
                "generatedAt": now_iso(),
            }
        except Exception as exc:
            errors.append(f"豆包: {exc}")
    elif not DOUBAO_API_KEY:
        errors.append("豆包 API Key 未配置")
    else:
        errors.append("豆包 Endpoint ID 未配置")

    return {"matchNo": match_no, "reports": reports, "errors": errors}

def get_ai_report(match_no: str) -> dict:
    match_no = normalize_match_no(match_no)
    data = load_data()
    match = next((m for m in data.get("matches", []) if normalize_match_no(m.get("matchNo", "")) == match_no), None)
    if not match:
        return {"matchNo": match_no, "reports": {}, "error": "比赛不存在"}

    reports_db = load_ai_reports()
    reports_db = purge_old_ai_reports(reports_db)
    existing = reports_db.get(match_no, {})

    result = {"matchNo": match_no, "reports": {}}
    if existing.get("deepseek", {}).get("content"):
        result["reports"]["deepseek"] = existing["deepseek"]
    if existing.get("doubao", {}).get("content"):
        result["reports"]["doubao"] = existing["doubao"]

    if not result["reports"]:
        result["pending"] = True
    return result

def trigger_ai_for_new_matches(data: dict) -> None:
    if not DEEPSEEK_API_KEY and not (DOUBAO_API_KEY and DOUBAO_ENDPOINT_ID):
        return
    reports_db = load_ai_reports()
    reports_db = purge_old_ai_reports(reports_db)
    any_new = False
    for match in data.get("matches", []):
        mn = normalize_match_no(match.get("matchNo", ""))
        if not mn:
            continue
        if mn not in reports_db:
            safe_print(f"[ai] 正在为比赛 {mn} 生成 AI 分析报告...")
            try:
                result = generate_ai_report_for_match(match)
                entry = {}
                for ai in ("deepseek", "doubao"):
                    if ai in result.get("reports", {}):
                        entry[ai] = result["reports"][ai]
                if entry:
                    reports_db[mn] = entry
                    any_new = True
                    safe_print(f"[ai] {mn} 分析完成: {list(entry.keys())}")
            except Exception as exc:
                safe_print(f"[ai] {mn} 分析失败: {exc}")
    if any_new:
        save_ai_reports(reports_db)
        safe_print(f"[ai] 本轮新增报告已保存")


def match_no_sort_key(match_no: str) -> tuple[int, str]:
    text = normalize_match_no(match_no)
    number_match = re.search(r"(\d+)", text)
    return (int(number_match.group(1)) if number_match else 999999, text)


def import_schedule(records: list[dict], data: dict, file_name: str) -> dict:
    by_no = {str(match.get("matchNo", "")).strip(): match for match in data.get("matches", []) if match.get("matchNo")}
    added = 0
    overwritten = 0
    errors: list[str] = []

    for row in records:
        try:
            match_no = normalize_match_no(row.get("match_no", ""))
            if not match_no:
                raise ValueError("竞彩编号为空")
            home, away = get_row_teams(row)
            kickoff_display = cell_text(row.get("kickoff", ""))
            kickoff = parse_kickoff(kickoff_display)
            day = day_payload(schedule_day_for_match(match_no, kickoff))
            existing = by_no.get(match_no)
            if existing:
                ensure_match_fields(existing)
            keep_score = existing and normalize_team(existing.get("home")) == normalize_team(home) and normalize_team(existing.get("away")) == normalize_team(away)

            match = {
                "matchNo": match_no,
                "league": str(row.get("league", "")).strip(),
                "round": str(row.get("round", "")).strip(),
                "home": home,
                "away": away,
                "homeBadge": team_badge(home),
                "awayBadge": team_badge(away),
                "kickoff": kickoff.isoformat(),
                "kickoffDisplay": kickoff_display,
                "dateKey": day["dateKey"],
                "dateLabel": day["dateLabel"],
                "dayNumber": day["dayNumber"],
                "dayOfWeek": day["dayOfWeek"],
                "time": day["time"],
                "fullScore": existing.get("fullScore", "") if keep_score else "",
                "halfScore": existing.get("halfScore", "") if keep_score else "",
                "homeScore": existing.get("homeScore", "") if keep_score else "",
                "awayScore": existing.get("awayScore", "") if keep_score else "",
                "apiMatchId": existing.get("apiMatchId", "") if keep_score else "",
                "homeTeamAliasMatched": existing.get("homeTeamAliasMatched", "") if keep_score else "",
                "awayTeamAliasMatched": existing.get("awayTeamAliasMatched", "") if keep_score else "",
                "leagueAliasMatched": existing.get("leagueAliasMatched", "") if keep_score else "",
                "matchStatus": existing.get("matchStatus", "未开始") if keep_score else "未开始",
                "currentScore": existing.get("currentScore", "") if keep_score else "",
                "penaltyScore": existing.get("penaltyScore", "") if keep_score else "",
                "extraTimeScore": existing.get("extraTimeScore", "") if keep_score else "",
                "scoreSource": existing.get("scoreSource", "pending") if keep_score else "pending",
                "manualHalfTimeScore": existing.get("manualHalfTimeScore", "") if keep_score else "",
                "manualFullTimeScore": existing.get("manualFullTimeScore", "") if keep_score else "",
                "apiHalfTimeScore": existing.get("apiHalfTimeScore", "") if keep_score else "",
                "apiFullTimeScore": existing.get("apiFullTimeScore", "") if keep_score else "",
                "manualLocked": existing.get("manualLocked", False) if keep_score else False,
                "needsManualCheck": existing.get("needsManualCheck", False) if keep_score else False,
                "lastApiState": existing.get("lastApiState", "") if keep_score else "",
                "apiUpdatedAt": existing.get("apiUpdatedAt", "") if keep_score else "",
                "manualUpdatedAt": existing.get("manualUpdatedAt", "") if keep_score else "",
                "lastUpdatedAt": now_iso(),
                "sourceFile": file_name,
                "updatedAt": now_iso(),
                "createdAt": existing.get("createdAt", now_iso()) if existing else now_iso(),
            }
            ensure_match_fields(match)

            if existing:
                existing.update(match)
                overwritten += 1
            else:
                data.setdefault("matches", []).append(match)
                by_no[match_no] = match
                added += 1
        except Exception as exc:
            errors.append(f"第{row.get('__row', '?')}行：{exc}")

    return {"added": added, "overwritten": overwritten, "errors": errors}


STATUS_MAP = {
    "match finished": "已结束",
    "match finished after extra time": "已结束",
    "match finished after penalties": "已结束",
    "first half": "上半场",
    "second half": "下半场",
    "half time": "半场结束",
    "halftime": "半场结束",
    "extra time": "加时/点球中",
    "penalty in progress": "加时/点球中",
    "not started": "未开始",
    "time to be defined": "未开始",
    "scheduled": "未开始",
    "match suspended": "异常",
    "match interrupted": "异常",
    "match abandoned": "异常",
    "match awarded": "异常",
    "match postponed": "延期",
    "match cancelled": "取消",
    "match canceled": "取消",
}
TERMINAL_STATUS = {"已结束", "延期", "取消", "异常"}
HALF_TIME_LABEL = "半场结束"
SECOND_HALF_LABELS = {"下半场", "加时/点球中", "已结束"}


def normalize_lookup(value) -> str:
    text = str(value or "").lower()
    return re.sub(r"[\s\.\-_'’`·・/\\()（）\[\]【】:：]+", "", text)


def parse_iso_datetime(value) -> datetime | None:
    if isinstance(value, (int, float)):
        timestamp = float(value)
        if timestamp > 10_000_000_000:
            timestamp /= 1000
        return datetime.fromtimestamp(timestamp, ZoneInfo(LOCAL_TIMEZONE)).replace(tzinfo=None)
    text = str(value or "").strip()
    if not text:
        return None
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    for fmt in ("%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            parsed = datetime.strptime(text, fmt)
            if parsed.tzinfo:
                parsed = parsed.astimezone(ZoneInfo(LOCAL_TIMEZONE)).replace(tzinfo=None)
            return parsed.replace(second=0, microsecond=0)
        except ValueError:
            continue
    try:
        parsed = datetime.fromisoformat(text)
        if parsed.tzinfo:
            parsed = parsed.astimezone(ZoneInfo(LOCAL_TIMEZONE)).replace(tzinfo=None)
        return parsed.replace(second=0, microsecond=0)
    except ValueError:
        return None


def kickoff_datetime(match: dict) -> datetime | None:
    try:
        return datetime.fromisoformat(str(match.get("kickoff", "")))
    except ValueError:
        return None


def nested_get(payload, paths: list[tuple]) -> object:
    for path in paths:
        current = payload
        for key in path:
            if isinstance(current, dict) and key in current:
                current = current[key]
            else:
                current = None
                break
        if current not in (None, ""):
            return current
    return ""


def first_nested_name(payload, paths: list[tuple]) -> str:
    value = nested_get(payload, paths)
    if isinstance(value, dict):
        return str(value.get("name") or value.get("displayName") or value.get("title") or value.get("shortName") or "")
    return str(value or "")


def extract_api_matches(payload) -> list[dict]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if not isinstance(payload, dict):
        return []
    for key in ("data", "matches", "results", "response"):
        value = payload.get(key)
        if isinstance(value, list):
            return [item for item in value if isinstance(item, dict)]
        if isinstance(value, dict):
            nested = extract_api_matches(value)
            if nested:
                return nested
    return []


def extract_api_match_id(item: dict) -> str:
    value = nested_get(item, [
        ("id",),
        ("matchId",),
        ("match_id",),
        ("fixture", "id"),
        ("event", "id"),
    ])
    return str(value or "")


def extract_api_home(item: dict) -> str:
    return first_nested_name(item, [
        ("homeTeam",),
        ("home_team",),
        ("home",),
        ("teams", "home"),
        ("participants", "home"),
        ("localteam",),
    ])


def extract_api_away(item: dict) -> str:
    return first_nested_name(item, [
        ("awayTeam",),
        ("away_team",),
        ("away",),
        ("teams", "away"),
        ("participants", "away"),
        ("visitorteam",),
    ])


def extract_api_league(item: dict) -> str:
    return first_nested_name(item, [
        ("league",),
        ("competition",),
        ("tournament",),
        ("season", "league"),
    ])


def extract_api_start(item: dict) -> datetime | None:
    value = nested_get(item, [
        ("date",),
        ("datetime",),
        ("startTime",),
        ("start_time",),
        ("kickoff",),
        ("fixture", "date"),
        ("time", "starting_at", "date_time"),
        ("time", "starting_at"),
    ])
    return parse_iso_datetime(value)


def extract_api_status_raw(item: dict) -> str:
    value = nested_get(item, [
        ("fixture", "status", "long"),
        ("state", "description"),
        ("status", "description"),
        ("status", "long"),
        ("status", "short"),
    ])
    return str(value or "")


def map_api_status(raw_status: str) -> str:
    key = normalize_lookup(raw_status)
    for source, label in STATUS_MAP.items():
        norm_source = normalize_lookup(source)
        if norm_source == key:
            return label
    for source, label in STATUS_MAP.items():
        norm_source = normalize_lookup(source)
        if norm_source in key or key in norm_source:
            return label
    return "异常" if raw_status else "未开始"


def score_pair_from_value(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, dict):
        home = nested_get(value, [("home",), ("homeScore",), ("home_score",), ("current", "home"), ("goals", "home")])
        away = nested_get(value, [("away",), ("awayScore",), ("away_score",), ("current", "away"), ("goals", "away")])
        if str(home).strip().isdigit() and str(away).strip().isdigit():
            return f"{int(home)}∶{int(away)}"
    parsed = parse_score_safe(str(value))
    return parsed[2] if parsed else ""


def extract_score_candidates(item: dict) -> dict[str, str]:
    current = score_pair_from_value(nested_get(item, [
        ("score", "fulltime"),
        ("goals",),
        ("state", "score", "current"),
        ("score", "current"),
        ("currentScore",),
    ]))
    half = score_pair_from_value(nested_get(item, [
        ("score", "halftime"),
        ("state", "score", "halftime"),
        ("state", "score", "halfTime"),
        ("halfTimeScore",),
        ("halftimeScore",),
    ]))
    full = score_pair_from_value(nested_get(item, [
        ("score", "fulltime"),
        ("state", "score", "fulltime"),
        ("scores", "fulltime"),
        ("fullTimeScore",),
    ]))
    penalty = score_pair_from_value(nested_get(item, [
        ("state", "score", "penalty"),
        ("status", "score", "penalty"),
        ("score", "penalty"),
        ("scores", "penalty"),
        ("penaltyScore",),
    ]))
    extra = score_pair_from_value(nested_get(item, [
        ("state", "score", "extratime"),
        ("state", "score", "extraTime"),
        ("status", "score", "extratime"),
        ("status", "score", "extraTime"),
        ("score", "extratime"),
        ("score", "extraTime"),
        ("scores", "extraTime"),
        ("extraTimeScore",),
    ]))
    return {"current": current, "half": half, "full": full, "penalty": penalty, "extra": extra}


def alias_set(name: str, aliases: dict) -> set[str]:
    values = [str(name or "")]
    values.extend(str(item) for item in aliases.get(name, []))
    return {normalize_lookup(item) for item in values if normalize_lookup(item)}


def alias_match(local_name: str, remote_name: str, aliases: dict) -> tuple[bool, str, float]:
    remote_key = normalize_lookup(remote_name)
    if not remote_key:
        return False, "", 0.0
    best_label = ""
    best_score = 0.0
    for candidate in alias_set(local_name, aliases):
        if remote_key == candidate or remote_key in candidate or candidate in remote_key:
            return True, remote_name, 1.0
        ratio = SequenceMatcher(None, candidate, remote_key).ratio()
        if ratio > best_score:
            best_score = ratio
            best_label = remote_name
    return best_score >= 0.86, best_label, best_score


def league_match(local_league: str, remote_league: str, aliases: dict) -> tuple[bool, str]:
    if not local_league or not remote_league:
        return True, ""
    ok, label, score = alias_match(local_league, remote_league, aliases)
    if ok:
        return True, label
    return score >= 0.72, label


def build_api_index(api_matches: list[dict]) -> dict[str, dict]:
    result = {}
    for item in api_matches:
        match_id = extract_api_match_id(item)
        if match_id:
            result[match_id] = item
    return result


def find_api_match(local_match: dict, api_matches: list[dict], team_aliases: dict, league_aliases: dict) -> tuple[dict | None, dict]:
    ensure_match_fields(local_match)
    local_id = str(local_match.get("apiMatchId") or "")
    if local_id:
        api_index = build_api_index(api_matches)
        if local_id in api_index:
            return api_index[local_id], {"method": "id", "candidates": 1}

    kickoff = kickoff_datetime(local_match)
    candidates = []
    for item in api_matches:
        api_home = extract_api_home(item)
        api_away = extract_api_away(item)
        api_league = extract_api_league(item)
        api_start = extract_api_start(item)
        home_ok, home_label, home_score = alias_match(local_match.get("home", ""), api_home, team_aliases)
        away_ok, away_label, away_score = alias_match(local_match.get("away", ""), api_away, team_aliases)
        league_ok, league_label = league_match(local_match.get("league", ""), api_league, league_aliases)
        swapped = False
        if not home_ok or not away_ok:
            home_ok_s, home_label_s, home_score_s = alias_match(local_match.get("home", ""), api_away, team_aliases)
            away_ok_s, away_label_s, away_score_s = alias_match(local_match.get("away", ""), api_home, team_aliases)
            if home_ok_s and away_ok_s:
                home_ok, home_label, home_score = True, home_label_s, home_score_s
                away_ok, away_label, away_score = True, away_label_s, away_score_s
                swapped = True
        if not home_ok or not away_ok:
            continue
        time_delta = 0
        if kickoff and api_start:
            time_delta = abs((kickoff - api_start).total_seconds()) / 60
            if time_delta > 30:
                continue
        score = home_score + away_score + (0.25 if league_ok else 0) + max(0, 0.3 - time_delta / 100)
        candidates.append({
            "item": item,
            "score": score,
            "homeAlias": home_label_s if swapped else home_label,
            "awayAlias": away_label_s if swapped else away_label,
            "leagueAlias": league_label,
            "swapped": swapped,
        })

    if len(candidates) == 1:
        candidate = candidates[0]
        return candidate["item"], {"method": "alias", "candidates": 1, **candidate}
    if len(candidates) > 1:
        candidates.sort(key=lambda item: item["score"], reverse=True)
        top = candidates[0]
        second = candidates[1]
        if top["score"] - second["score"] >= 0.35:
            return top["item"], {"method": "alias", "candidates": len(candidates), **top}
        return None, {"method": "ambiguous", "candidates": len(candidates)}
    return None, {"method": "none", "candidates": 0}


def update_match_from_api(local_match: dict, api_match: dict, match_info: dict) -> bool:
    ensure_match_fields(local_match)
    changed = False
    api_id = extract_api_match_id(api_match)
    if api_id and str(local_match.get("apiMatchId") or "") != api_id:
        local_match["apiMatchId"] = api_id
        changed = True
    for field, info_key in (
        ("homeTeamAliasMatched", "homeAlias"),
        ("awayTeamAliasMatched", "awayAlias"),
        ("leagueAliasMatched", "leagueAlias"),
    ):
        value = str(match_info.get(info_key) or "")
        if value and local_match.get(field) != value:
            local_match[field] = value
            changed = True

    status_raw = extract_api_status_raw(api_match)
    status_label = map_api_status(status_raw)
    scores = extract_score_candidates(api_match)

    fulltime = scores["full"]
    halftime = scores["half"]

    has_manual_score = bool(local_match.get("manualHalfTimeScore") or local_match.get("manualFullTimeScore"))
    had_api_score = bool(local_match.get("apiHalfTimeScore") or local_match.get("apiFullTimeScore") or local_match.get("currentScore"))

    if status_label != local_match.get("matchStatus"):
        local_match["matchStatus"] = status_label
        changed = True
    if status_raw and status_raw != local_match.get("lastApiState"):
        local_match["lastApiState"] = status_raw
        changed = True

    if fulltime and parse_score_safe(fulltime):
        if local_match.get("currentScore") != fulltime:
            local_match["currentScore"] = fulltime
            changed = True

    if halftime:
        if local_match.get("apiHalfTimeScore") != halftime:
            local_match["apiHalfTimeScore"] = halftime
            changed = True

    if status_label == HALF_TIME_LABEL and fulltime and parse_score_safe(fulltime):
        if local_match.get("apiHalfTimeScore") != fulltime:
            local_match["apiHalfTimeScore"] = fulltime
            changed = True

    if fulltime and parse_score_safe(fulltime):
        if local_match.get("apiFullTimeScore") != fulltime:
            local_match["apiFullTimeScore"] = fulltime
            changed = True

    if scores["penalty"]:
        local_match["penaltyScore"] = scores["penalty"]
    if scores["extra"]:
        local_match["extraTimeScore"] = scores["extra"]

    if changed or fulltime:
        local_match["apiUpdatedAt"] = now_iso()
        local_match["updatedAt"] = now_iso()
        local_match["lastUpdatedAt"] = now_iso()
        if has_manual_score and (local_match.get("apiFullTimeScore") or local_match.get("apiHalfTimeScore")):
            local_match["scoreSource"] = "api_over_manual"
        elif local_match.get("apiFullTimeScore") or local_match.get("apiHalfTimeScore") or local_match.get("currentScore"):
            local_match["scoreSource"] = "api"
        elif not had_api_score:
            local_match["scoreSource"] = local_match.get("scoreSource") or "pending"
        apply_effective_score_fields(local_match)
    return changed


def update_usage_after_request(usage: dict, *, status: int, remaining_header: str = "", ok: bool = True, error: str = "") -> dict:
    usage = normalize_api_usage(usage)
    usage["used"] = min(API_DAILY_LIMIT, int(usage.get("used", 0)) + 1)
    usage["lastRequestAt"] = now_iso()
    if remaining_header and str(remaining_header).isdigit():
        usage["remaining"] = max(0, int(remaining_header))
    else:
        usage["remaining"] = max(0, API_DAILY_LIMIT - int(usage.get("used", 0)))
    if ok:
        usage["lastSuccessAt"] = now_iso()
        usage["lastError"] = ""
        usage["consecutiveFailures"] = 0
        usage["pauseUntil"] = None
    else:
        usage["lastError"] = error
        usage["consecutiveFailures"] = int(usage.get("consecutiveFailures", 0)) + 1
        if status == 429 or usage["remaining"] <= 0:
            usage["remaining"] = 0
        if usage["consecutiveFailures"] >= 3:
            usage["pauseUntil"] = (datetime.now() + timedelta(minutes=5)).isoformat(timespec="seconds")
    usage["lastUpdatedAt"] = now_iso()
    save_api_usage(usage)
    return usage


def apifootball_quota_available(force: bool = False) -> tuple[bool, dict, str]:
    usage = load_api_usage()
    if not API_FOOTBALL_API_KEY:
        return False, usage, "未配置 API_FOOTBALL_API_KEY"
    if usage.get("pauseUntil") and not force:
        pause_until = parse_iso_datetime(usage["pauseUntil"])
        if pause_until and pause_until > datetime.now():
            return False, usage, "连续失败后暂停中"
    if int(usage.get("remaining", 0)) <= 0:
        return False, usage, "API-Football 今日额度已用完"
    if int(usage.get("remaining", 0)) <= SCORE_POLL_MIN_REMAINING and not force:
        return False, usage, "剩余额度进入省流模式"
    return True, usage, ""


def _fetch_apifootball_page(api_date: date) -> list[dict]:
    query = urlencode({"date": api_date.isoformat()})
    url = f"{API_FOOTBALL_BASE_URL}/fixtures?{query}"
    request = urllib.request.Request(
        url,
        headers={
            "Accept": "application/json",
            "User-Agent": "QTC-APIFootballScoreUpdater/1.0",
            "x-apisports-key": API_FOOTBALL_API_KEY,
        },
    )
    with urllib.request.urlopen(request, timeout=10) as response:
        body = response.read()
        payload = json.loads(body.decode("utf-8"))
        # 检查 JSON 响应中的错误（API 可能返回 HTTP 200 但包含限流等错误）
        errors = payload.get("errors")
        if isinstance(errors, dict) and errors:
            messages = []
            for v in errors.values():
                if isinstance(v, str):
                    messages.append(v)
                elif isinstance(v, list):
                    messages.extend(str(x) for x in v)
                elif isinstance(v, dict):
                    messages.extend(str(x) for x in v.values())
            raise OSError("; ".join(messages) if messages else str(errors))
        return extract_api_matches(payload)


def fetch_apifootball_matches(api_date: date, force: bool = False) -> tuple[list[dict], dict]:
    ok, usage, reason = apifootball_quota_available(force=force)
    if not ok:
        raise RuntimeError(reason)

    started = now_iso()
    query = urlencode({"date": api_date.isoformat()})
    try:
        matches = _fetch_apifootball_page(api_date)
    except urllib.error.HTTPError as exc:
        remaining = exc.headers.get("x-ratelimit-requests-remaining", "") if exc.headers else ""
        message = exc.read().decode("utf-8", errors="replace")[:500]
        usage = update_usage_after_request(usage, status=exc.code, remaining_header=remaining, ok=False, error=message)
        url = f"{API_FOOTBALL_BASE_URL}/fixtures?{query}".replace(API_FOOTBALL_API_KEY, "***")
        append_api_log({"requestAt": started, "url": url, "status": exc.code, "error": message, "remaining": usage.get("remaining")})
        raise RuntimeError(f"API-Football 请求失败：HTTP {exc.code}") from exc
    except (OSError, json.JSONDecodeError) as exc:
        usage = update_usage_after_request(usage, status=0, ok=False, error=str(exc))
        url = f"{API_FOOTBALL_BASE_URL}/fixtures?{query}".replace(API_FOOTBALL_API_KEY, "***")
        append_api_log({"requestAt": started, "url": url, "status": 0, "error": str(exc), "remaining": usage.get("remaining")})
        raise RuntimeError(f"API-Football 请求失败：{exc}") from exc

    usage = update_usage_after_request(load_api_usage(), status=200, ok=True)
    url = f"{API_FOOTBALL_BASE_URL}/fixtures?{query}".replace(API_FOOTBALL_API_KEY, "***")
    append_api_log({
        "requestAt": started,
        "url": url,
        "status": 200,
        "remaining": usage.get("remaining"),
        "matchCount": len(matches),
        "updatedMatches": [],
    })
    return matches, usage


def api_dates_for_matches(matches: list[dict], requested: str = "", *, active_only: bool = False) -> list[date]:
    if requested:
        return [date.fromisoformat(requested)]
    result: list[date] = []
    for match in matches:
        ensure_match_fields(match)
        kickoff = kickoff_datetime(match)
        if not kickoff:
            continue
        match_date = kickoff.date()
        # 已结束的比赛无需再查
        if match.get("matchStatus") in TERMINAL_STATUS:
            continue
        result.append(match_date)
    # 按日期倒序，优先查最新的比赛，最多 5 个日期以控制 API 配额
    return sorted(set(result), reverse=True)[:5]


def run_apifootball_update(data: dict, *, requested_date: str = "", force: bool = False, active_only: bool = False) -> dict:
    team_aliases = load_aliases(TEAM_ALIASES_FILE, DEFAULT_TEAM_ALIASES)
    league_aliases = load_aliases(LEAGUE_ALIASES_FILE, DEFAULT_LEAGUE_ALIASES)
    local_matches = data.get("matches", [])
    updated = 0
    matched = 0
    errors: list[str] = []
    touched: list[str] = []

    # 按日期批量查询 API-Football（一次查询返回当日所有比赛，高效省配额）
    api_dates = api_dates_for_matches(local_matches, requested_date, active_only=active_only)
    for api_date in api_dates:
        try:
            api_matches, usage = fetch_apifootball_matches(api_date, force=force)
        except Exception as exc:
            errors.append(str(exc))
            continue
        for local_match in local_matches:
            ensure_match_fields(local_match)
            kickoff = kickoff_datetime(local_match)
            if not kickoff or kickoff.date() != api_date:
                continue
            # 已结束的比赛跳过
            if local_match.get("matchStatus") in TERMINAL_STATUS:
                continue
            api_match, info = find_api_match(local_match, api_matches, team_aliases, league_aliases)
            if not api_match:
                continue
            matched += 1
            if update_match_from_api(local_match, api_match, info):
                updated += 1
            touched.append(str(local_match.get("matchNo") or ""))

    usage = load_api_usage()
    data["lastApiUpdate"] = {
        "status": "success" if not errors else "partial" if updated or matched else "failed",
        "updated": updated,
        "matched": matched,
        "errors": errors[:20],
        "updatedMatches": touched[:50],
        "updatedAt": now_iso(),
        "apiUsage": usage,
    }
    append_api_log({
        "kind": "update-summary",
        "updated": updated,
        "matched": matched,
        "errors": errors[:20],
        "updatedMatches": touched[:50],
        "remaining": usage.get("remaining"),
    })
    return data["lastApiUpdate"]


def score_scheduler_forever() -> None:
    """每 30 分钟轮询 API-Football 一次，获取半场/全场比分和比赛状态"""
    while True:
        try:
            with DATA_LOCK:
                data = load_data()
                result = run_apifootball_update(data, force=False, active_only=False)
                save_data(data)
            if result.get("updated") or result.get("matched"):
                broadcast({"kind": "updated", "result": result})
            safe_print(f"[apifootball] 本轮完成：更新 {result.get('updated', 0)}，匹配 {result.get('matched', 0)}，错误 {len(result.get('errors', []))}")
        except Exception as exc:
            safe_print(f"[apifootball] {exc}")
        time.sleep(1800)  # 30 分钟


def import_workbook(path: Path, data: dict) -> dict:
    _, records, _ = read_excel(path)
    result = import_schedule(records, data, path.name)
    message = f"赛程导入完成：新增 {result['added']}，覆盖 {result['overwritten']}，失败 {len(result['errors'])}"

    purged = purge_old_matches(data)
    if purged:
        message += f"，清理过期 {purged}"

    status = "failed" if records and result["errors"] and not any(result.get(key, 0) for key in ("added", "overwritten", "updated")) else "success"
    data["lastImport"] = {
        "status": status,
        "type": "schedule",
        "file": path.name,
        "message": message,
        "errors": result["errors"][:30],
        "importedAt": now_iso(),
    }
    return data["lastImport"]


def build_schedule_payload(data: dict) -> dict:
    # Make a working copy so we don't modify the original data during read
    # Purge only happens during import and scheduled cleanup, not on every API request
    matches = sorted(
        data.get("matches", []),
        key=lambda item: (item.get("dateKey", ""), *match_no_sort_key(item.get("matchNo", "")), item.get("time", "")),
    )
    grouped: dict[str, dict] = {}
    for match in matches:
        ensure_match_fields(match)
        date_key = match["dateKey"]
        grouped.setdefault(date_key, {
            "dateKey": date_key,
            "dateLabel": match["dateLabel"],
            "dayNumber": match["dayNumber"],
            "dayOfWeek": match["dayOfWeek"],
            "matches": [],
        })
        grouped[date_key]["matches"].append(match)

    return {"days": list(grouped.values()), "lastImport": data.get("lastImport")}


def latest_image(kind: str) -> Path | None:
    folder = DAILY_IMAGE_DIRS.get(kind)
    if not folder:
        return None
    images = [
        path for path in folder.iterdir()
        if path.is_file() and path.suffix.lower() in IMAGE_EXTENSIONS and not path.name.startswith("~$")
    ]
    if not images:
        return None
    return max(images, key=lambda path: path.stat().st_mtime)


def valid_team_logo_source(src: str) -> bool:
    parsed = urlparse(src)
    host = (parsed.hostname or "").lower()
    return parsed.scheme == "https" and host in TEAM_LOGO_ALLOWED_HOSTS and not parsed.username and not parsed.password


def team_logo_cache_path(src: str) -> Path:
    parsed = urlparse(src)
    suffix = Path(parsed.path).suffix.lower()
    if suffix not in TEAM_LOGO_EXTENSIONS:
        suffix = ".img"
    digest = hashlib.sha256(src.encode("utf-8")).hexdigest()
    return TEAM_LOGO_CACHE_DIR / f"{digest}{suffix}"


def team_logo_content_type(path: Path) -> str:
    if path.suffix.lower() == ".svg":
        return "image/svg+xml"
    return mimetypes.guess_type(path.name)[0] or "application/octet-stream"


def fetch_team_logo(src: str, target: Path) -> None:
    request = urllib.request.Request(
        src,
        headers={
            "Accept": "image/avif,image/webp,image/svg+xml,image/png,image/*,*/*;q=0.8",
            "User-Agent": "QTC-TeamLogoCache/1.0",
        },
    )
    with urllib.request.urlopen(request, timeout=8) as response:
        body = response.read(TEAM_LOGO_MAX_BYTES + 1)
        if len(body) > TEAM_LOGO_MAX_BYTES:
            raise ValueError("team logo too large")
        content_type = response.headers.get_content_type()
        if not (content_type.startswith("image/") or target.suffix.lower() == ".svg"):
            raise ValueError(f"invalid content type: {content_type}")

    target.parent.mkdir(parents=True, exist_ok=True)
    temp = target.with_suffix(target.suffix + ".tmp")
    temp.write_bytes(body)
    temp.replace(target)


def safe_move(path: Path, target_dir: Path) -> Path:
    target_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target = target_dir / f"{stamp}_{path.name}"
    counter = 1
    while target.exists():
        target = target_dir / f"{stamp}_{counter}_{path.name}"
        counter += 1
    shutil.move(str(path), str(target))
    return target


def write_failure_note(target: Path, message: str) -> None:
    note = target.with_suffix(target.suffix + ".错误说明.txt")
    note.write_text(message, encoding="utf-8")


def upload_signature(path: Path) -> tuple[int, int]:
    stat = path.stat()
    return stat.st_size, stat.st_mtime_ns


def excel_lock_path(path: Path) -> Path:
    return path.with_name(f"~${path.name}")


def has_excel_lock(path: Path) -> bool:
    return excel_lock_path(path).exists()


def wait_until_stable(path: Path, timeout: int = 20) -> None:
    last_signature = (-1, -1)
    stable_ticks = 0
    start = time.time()
    while time.time() - start < timeout:
        if not path.exists():
            raise FileNotFoundError(path)
        if has_excel_lock(path):
            stable_ticks = 0
            time.sleep(0.6)
            continue

        signature = upload_signature(path)
        if signature == last_signature and signature[0] > 0:
            stable_ticks += 1
            if stable_ticks >= 5:
                return
        else:
            stable_ticks = 0
            last_signature = signature
        time.sleep(0.6)
    raise TimeoutError(f"等待文件写入完成超时：{path.name}")


def broadcast(payload: dict) -> None:
    with CLIENTS_LOCK:
        clients = list(CLIENTS)
    for client in clients:
        try:
            client.put_nowait(payload)
        except queue.Full:
            pass


def process_upload_file(path: Path) -> None:
    if path.name.startswith("~$"):
        return
    if path.suffix.lower() not in {".xlsx", ".xls", ".xlsm"}:
        return
    if has_excel_lock(path):
        return

    resolved = path.resolve()
    if resolved in IN_PROGRESS:
        return
    IN_PROGRESS.add(resolved)

    try:
        wait_until_stable(path)
        if has_excel_lock(path):
            return
        signature = upload_signature(path)
        if PROCESSED_UPLOAD_SIGNATURES.get(path.name) == signature:
            return
        with DATA_LOCK:
            data = load_data()
            try:
                result = import_workbook(path, data)
                if API_FOOTBALL_API_KEY:
                    try:
                        data["lastApiUpdate"] = run_apifootball_update(data, force=True)
                    except Exception as exc:
                        result.setdefault("errors", []).append(f"API-Football 初次匹配失败：{exc}")
                save_data(data)
                threading.Thread(target=trigger_ai_for_new_matches, args=(data,), name="ai-trigger", daemon=True).start()
                target_dir = FAILED_DIR if result.get("status") == "failed" else IMPORTED_DIR
                moved_to = safe_move(path, target_dir)
                result["storedAs"] = moved_to.name
                if result.get("status") == "failed":
                    write_failure_note(moved_to, "\n".join(result.get("errors", [])) or result.get("message", "导入失败"))
                PROCESSED_UPLOAD_SIGNATURES[path.name] = signature
                broadcast({"kind": "updated", "result": result})
            except Exception as exc:
                data["lastImport"] = {
                    "status": "failed",
                    "type": "unknown",
                    "file": path.name,
                    "message": f"导入失败：{exc}",
                    "errors": [str(exc)],
                    "importedAt": now_iso(),
                }
                save_data(data)
                moved_to = safe_move(path, FAILED_DIR)
                write_failure_note(moved_to, str(exc))
                PROCESSED_UPLOAD_SIGNATURES[path.name] = signature
                broadcast({"kind": "updated", "result": data["lastImport"]})
    finally:
        IN_PROGRESS.discard(resolved)


def scan_uploads_forever() -> None:
    while True:
        try:
            ensure_dirs()
            for path in sorted(UPLOAD_DIR.iterdir(), key=lambda item: item.stat().st_mtime):
                if path.is_file():
                    process_upload_file(path)
        except Exception as exc:
            print(f"[watcher] {exc}", flush=True)
        time.sleep(SCAN_SECONDS)


ALLOWED_STATIC = {"index.html", "styles.css", "script.js"}


class AppHandler(BaseHTTPRequestHandler):
    server_version = "QTCSchedule/1.0"

    def log_message(self, fmt: str, *args) -> None:
        print("[%s] %s" % (self.log_date_time_string(), fmt % args), flush=True)

    def send_json(self, payload: dict, status: HTTPStatus = HTTPStatus.OK) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def read_json_body(self) -> dict:
        length = int(self.headers.get("Content-Length", "0") or "0")
        if length <= 0:
            return {}
        body = self.rfile.read(length)
        try:
            return json.loads(body.decode("utf-8"))
        except json.JSONDecodeError:
            raise ValueError("请求体必须是 JSON")

    def is_admin_request(self) -> bool:
        token = self.headers.get("X-Admin-Token", "")
        auth = self.headers.get("Authorization", "")
        if auth.lower().startswith("bearer "):
            token = auth[7:].strip()
        if not ADMIN_TOKEN:
            return self.client_address[0] in {"127.0.0.1", "::1", "localhost"}
        return token == ADMIN_TOKEN

    def require_admin(self) -> bool:
        if self.is_admin_request():
            return True
        self.send_json({"ok": False, "message": "管理员接口需要 X-Admin-Token 或 Bearer Token"}, HTTPStatus.UNAUTHORIZED)
        return False

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        route = parsed.path

        if route == "/api/schedule":
            with DATA_LOCK:
                data = load_data()
                payload = build_schedule_payload(data)
            self.send_json(payload)
            return

        if route == "/api/matches/today":
            today = date.today().isoformat()
            with DATA_LOCK:
                data = load_data()
                payload = {
                    "date": today,
                    "matches": [
                        match for match in data.get("matches", [])
                        if (kickoff_datetime(match) and kickoff_datetime(match).date().isoformat() == today)
                    ],
                    "lastApiUpdate": data.get("lastApiUpdate"),
                }
            self.send_json(payload)
            return

        if route == "/api/update-status":
            with DATA_LOCK:
                data = load_data()
                payload = {
                    "apiUsage": load_api_usage(),
                    "lastApiUpdate": data.get("lastApiUpdate"),
                    "lastImport": data.get("lastImport"),
                    "configured": bool(API_FOOTBALL_API_KEY),
                }
            self.send_json(payload)
            return

        if route == "/api/image/latest":
            query = {key: values[-1] for key, values in parse_qs(parsed.query).items() if values}
            kind = query.get("kind", "")
            image = latest_image(kind)
            if not image:
                self.send_json({"found": False, "kind": kind})
                return
            self.send_json({
                "found": True,
                "kind": kind,
                "name": image.name,
                "url": f"/daily-image/{kind}/{image.name}?v={int(image.stat().st_mtime)}",
            })
            return

        if route == "/api/team-logo":
            self.serve_team_logo(parsed.query)
            return

        if route == "/api/ai-report":
            query = {key: values[-1] for key, values in parse_qs(parsed.query).items() if values}
            match_no = query.get("matchNo", "")
            if not match_no:
                self.send_json({"ok": False, "message": "缺少 matchNo 参数"}, HTTPStatus.BAD_REQUEST)
                return
            result = get_ai_report(match_no)
            self.send_json({"ok": True, **result})
            return

        if route == "/api/events":
            self.handle_events()
            return

        if route.startswith("/daily-image/"):
            self.serve_daily_image(route)
            return

        self.serve_static(route)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        route = parsed.path
        if not route.startswith("/admin/"):
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        if not self.require_admin():
            return

        try:
            payload = self.read_json_body()
            if route == "/admin/import-matches":
                records = payload.get("matches") or payload.get("records") or []
                if not isinstance(records, list):
                    raise ValueError("matches 必须是数组")
                normalized_records = []
                for index, item in enumerate(records, start=1):
                    normalized_records.append({
                        "__row": index,
                        "match_no": item.get("match_no") or item.get("matchNo") or item.get("jcNo") or item.get("竞彩编号"),
                        "league": item.get("league") or item.get("leagueName") or item.get("赛事种类") or item.get("赛事"),
                        "home": item.get("home") or item.get("homeTeamName") or item.get("主队"),
                        "away": item.get("away") or item.get("awayTeamName") or item.get("客队"),
                        "teams": item.get("teams") or item.get("双方队伍"),
                        "kickoff": item.get("kickoff") or item.get("startTime") or item.get("开赛日期") or item.get("开赛时间"),
                    })
                with DATA_LOCK:
                    data = load_data()
                    result = import_schedule(normalized_records, data, "admin-json")
                    data["lastImport"] = {"status": "success", "type": "schedule", "message": "后台导入赛程完成", "errors": result["errors"], "importedAt": now_iso()}
                    save_data(data)
                broadcast({"kind": "updated", "result": result})
                self.send_json({"ok": True, "result": result})
                return

            if route in {"/admin/match-api-football", "/admin/update-scores"}:
                requested_date = str(payload.get("date") or "")
                with DATA_LOCK:
                    data = load_data()
                    result = run_apifootball_update(data, requested_date=requested_date, force=True)
                    save_data(data)
                broadcast({"kind": "updated", "result": result})
                self.send_json({"ok": True, "result": result})
                return

            if route == "/admin/bind-api-match-id":
                match_no = normalize_match_no(payload.get("matchNo") or payload.get("jcNo") or payload.get("竞彩编号") or "")
                api_match_id = str(payload.get("apiMatchId") or payload.get("matchId") or "").strip()
                if not match_no or not api_match_id:
                    raise ValueError("必须提供 matchNo 和 apiMatchId")
                with DATA_LOCK:
                    data = load_data()
                    target = next((item for item in data.get("matches", []) if normalize_match_no(item.get("matchNo", "")) == match_no), None)
                    if not target:
                        raise ValueError(f"找不到比赛：{match_no}")
                    ensure_match_fields(target)
                    target["apiMatchId"] = api_match_id
                    target["lastUpdatedAt"] = now_iso()
                    save_data(data)
                broadcast({"kind": "updated", "result": {"matchNo": match_no}})
                self.send_json({"ok": True, "matchNo": match_no, "apiMatchId": api_match_id})
                return

            if route == "/admin/edit-score":
                match_no = normalize_match_no(payload.get("matchNo") or payload.get("jcNo") or payload.get("竞彩编号") or "")
                if not match_no:
                    raise ValueError("必须提供 matchNo")
                half_score = str(payload.get("halfScore") or payload.get("halfTimeScore") or payload.get("半场比分") or "").strip()
                full_score = str(payload.get("fullScore") or payload.get("fullTimeScore") or payload.get("全场比分") or "").strip()
                with DATA_LOCK:
                    data = load_data()
                    target = next((item for item in data.get("matches", []) if normalize_match_no(item.get("matchNo", "")) == match_no), None)
                    if not target:
                        raise ValueError(f"找不到比赛：{match_no}")
                    ensure_match_fields(target)
                    if half_score:
                        _, _, target["manualHalfTimeScore"] = parse_score(half_score)
                    if full_score:
                        _, _, target["manualFullTimeScore"] = parse_score(full_score)
                    target["manualLocked"] = bool(payload.get("manualLocked", target.get("manualLocked", False)))
                    target["manualUpdatedAt"] = now_iso()
                    if not (target.get("apiFullTimeScore") or target.get("apiHalfTimeScore") or target.get("currentScore")):
                        target["scoreSource"] = "manual"
                    else:
                        target["scoreSource"] = "api_over_manual"
                    target["lastUpdatedAt"] = now_iso()
                    target["updatedAt"] = now_iso()
                    apply_effective_score_fields(target)
                    save_data(data)
                broadcast({"kind": "updated", "result": {"matchNo": match_no}})
                self.send_json({"ok": True, "matchNo": match_no})
                return

            self.send_error(HTTPStatus.NOT_FOUND)
        except Exception as exc:
            self.send_json({"ok": False, "message": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_events(self) -> None:
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "text/event-stream; charset=utf-8")
        self.send_header("Cache-Control", "no-cache")
        self.send_header("Connection", "keep-alive")
        self.end_headers()

        client_queue: queue.Queue[dict] = queue.Queue(maxsize=8)
        with CLIENTS_LOCK:
            CLIENTS.append(client_queue)

        try:
            self.wfile.write(b": connected\n\n")
            self.wfile.flush()
            while True:
                try:
                    payload = client_queue.get(timeout=15)
                    line = "event: updated\ndata: " + json.dumps(payload, ensure_ascii=False) + "\n\n"
                except queue.Empty:
                    line = ": keep-alive\n\n"
                self.wfile.write(line.encode("utf-8"))
                self.wfile.flush()
        except (BrokenPipeError, ConnectionResetError, TimeoutError):
            pass
        finally:
            with CLIENTS_LOCK:
                if client_queue in CLIENTS:
                    CLIENTS.remove(client_queue)

    def serve_static(self, route: str) -> None:
        if route in {"", "/"}:
            route = "/index.html"
        relative = unquote(route).lstrip("/")
        if relative not in ALLOWED_STATIC:
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        candidate = (ROOT / relative).resolve()
        if not candidate.exists() or not candidate.is_file():
            self.send_error(HTTPStatus.NOT_FOUND)
            return

        content_type = mimetypes.guess_type(candidate.name)[0] or "application/octet-stream"
        body = candidate.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type + ("; charset=utf-8" if content_type.startswith("text/") or candidate.suffix in {".js", ".css"} else ""))
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def serve_daily_image(self, route: str) -> None:
        parts = [unquote(part) for part in route.split("/") if part]
        if len(parts) != 3:
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        _, kind, name = parts
        folder = DAILY_IMAGE_DIRS.get(kind)
        if not folder:
            self.send_error(HTTPStatus.NOT_FOUND)
            return

        candidate = (folder / name).resolve()
        if folder.resolve() not in candidate.parents or not candidate.is_file():
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        if candidate.suffix.lower() not in IMAGE_EXTENSIONS:
            self.send_error(HTTPStatus.FORBIDDEN)
            return

        content_type = mimetypes.guess_type(candidate.name)[0] or "application/octet-stream"
        body = candidate.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Cache-Control", "public, max-age=31536000, immutable")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def serve_team_logo(self, query_string: str) -> None:
        query = parse_qs(query_string)
        src = query.get("src", [""])[0]
        if not src or not valid_team_logo_source(src):
            self.send_error(HTTPStatus.BAD_REQUEST)
            return

        cache_path = team_logo_cache_path(src)
        if not cache_path.exists():
            try:
                fetch_team_logo(src, cache_path)
            except (OSError, ValueError, urllib.error.URLError):
                self.send_error(HTTPStatus.NOT_FOUND)
                return

        body = cache_path.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", team_logo_content_type(cache_path))
        self.send_header("Cache-Control", "public, max-age=31536000, immutable")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


TEAM_FLAG_CODES: dict[str, str] = {
    "中国": "cn", "中国香港": "hk", "中国澳门": "mo", "中华台北": "tw",
    "日本": "jp", "韩国": "kr", "朝鲜": "kp",
    "澳大利亚": "au", "新西兰": "nz",
    "新加坡": "sg", "泰国": "th", "越南": "vn", "马来西亚": "my",
    "印尼": "id", "印度尼西亚": "id",
    "菲律宾": "ph", "缅甸": "mm", "柬埔寨": "kh", "老挝": "la",
    "印度": "in", "巴基斯坦": "pk", "孟加拉国": "bd",
    "伊朗": "ir", "伊拉克": "iq",
    "沙特": "sa", "沙特阿拉伯": "sa", "卡塔尔": "qa",
    "阿联酋": "ae", "阿曼": "om", "科威特": "kw", "巴林": "bh",
    "约旦": "jo", "叙利亚": "sy", "黎巴嫩": "lb",
    "巴勒斯坦": "ps", "以色列": "il", "也门": "ye",
    "乌兹别克": "uz", "乌兹别克斯坦": "uz",
    "哈萨克": "kz", "哈萨克斯坦": "kz",
    "吉尔吉斯": "kg", "吉尔吉斯斯坦": "kg",
    "塔吉克": "tj", "塔吉克斯坦": "tj",
    "土库曼": "tm", "土库曼斯坦": "tm",
    "英格兰": "gb-eng", "苏格兰": "gb-sct", "威尔士": "gb-wls", "北爱尔兰": "gb-nir",
    "爱尔兰": "ie", "法国": "fr", "德国": "de", "意大利": "it", "西班牙": "es",
    "葡萄牙": "pt", "荷兰": "nl", "比利时": "be", "瑞士": "ch",
    "奥地利": "at", "丹麦": "dk", "瑞典": "se", "挪威": "no",
    "芬兰": "fi", "冰岛": "is", "波兰": "pl", "捷克": "cz",
    "斯洛伐克": "sk", "匈牙利": "hu", "罗马尼亚": "ro",
    "保加利亚": "bg", "克罗地亚": "hr", "塞尔维亚": "rs",
    "黑山": "me", "波黑": "ba", "斯洛文尼亚": "si",
    "阿尔巴尼亚": "al", "北马其顿": "mk", "希腊": "gr",
    "土耳其": "tr", "乌克兰": "ua", "俄罗斯": "ru",
    "白俄罗斯": "by", "格鲁吉亚": "ge", "亚美尼亚": "am",
    "阿塞拜疆": "az", "立陶宛": "lt", "拉脱维亚": "lv",
    "爱沙尼亚": "ee", "摩尔多瓦": "md", "卢森堡": "lu",
    "巴西": "br", "阿根廷": "ar", "乌拉圭": "uy",
    "哥伦比亚": "co", "智利": "cl", "秘鲁": "pe",
    "巴拉圭": "py", "玻利维亚": "bo", "厄瓜多尔": "ec",
    "委内瑞拉": "ve", "美国": "us", "加拿大": "ca",
    "墨西哥": "mx", "巴拿马": "pa", "洪都拉斯": "hn",
    "哥斯达黎加": "cr", "萨尔瓦多": "sv", "危地马拉": "gt",
    "牙买加": "jm", "海地": "ht", "古巴": "cu",
    "埃及": "eg", "摩洛哥": "ma", "突尼斯": "tn",
    "阿尔及利亚": "dz", "尼日利亚": "ng", "加纳": "gh",
    "喀麦隆": "cm", "科特迪瓦": "ci", "塞内加尔": "sn",
    "南非": "za", "马里": "ml", "几内亚": "gn",
    "刚果": "cg", "民主刚果": "cd", "肯尼亚": "ke",
}

FLAG_CDN_BASE = "https://jsd.onmicrosoft.cn/npm/svg-country-flags@1.2.10/svg"


def flag_logo_url(team_name: str) -> str | None:
    code = TEAM_FLAG_CODES.get(team_name.strip())
    if not code:
        return None
    return f"{FLAG_CDN_BASE}/{code.lower()}.svg"


def precache_team_logos_forever() -> None:
    from concurrent.futures import ThreadPoolExecutor
    time.sleep(2)
    while True:
        try:
            with DATA_LOCK:
                data = load_data()
            teams: set[str] = set()
            for m in data.get("matches", []):
                for key in ("home", "away"):
                    name = (m.get(key) or "").strip()
                    if name:
                        teams.add(name)

            urls: list[str] = []
            for team in teams:
                url = flag_logo_url(team)
                if url:
                    urls.append(url)

            missing = [u for u in urls if not team_logo_cache_path(u).exists()]
            if missing:
                cached = 0
                with ThreadPoolExecutor(max_workers=3) as pool:
                    futures = {pool.submit(fetch_team_logo, u, team_logo_cache_path(u)): u for u in missing}
                    for f in futures:
                        try:
                            f.result(timeout=15)
                            cached += 1
                        except Exception:
                            pass
                safe_print(f"[logo-precache] downloaded {cached} new flag(s) to cache")
        except Exception:
            pass
        time.sleep(300)


def main() -> None:
    ensure_dirs()
    with DATA_LOCK:
        data = load_data()
        purge_old_matches(data)
        save_data(data)

    logo_precache = threading.Thread(target=precache_team_logos_forever, name="logo-precache", daemon=True)
    logo_precache.start()
    watcher = threading.Thread(target=scan_uploads_forever, name="excel-watcher", daemon=True)
    watcher.start()
    score_watcher = threading.Thread(target=score_scheduler_forever, name="apifootball-score-watcher", daemon=True)
    score_watcher.start()

    httpd = ThreadingHTTPServer((HOST, PORT), AppHandler)
    safe_print(f"Serving schedule app at http://localhost:{PORT}/")
    safe_print(f"Excel upload folder: {UPLOAD_DIR}")
    httpd.serve_forever()


if __name__ == "__main__":
    main()
